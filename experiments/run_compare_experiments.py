import os
import sys
import time
import json
import statistics

from openpyxl import Workbook
from openpyxl.styles import Font

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, PROJECT_ROOT)

from src.problem.instance_generator import load_instance_from_json
from src.algorithms.baseline_ga import BaselineGA
from src.algorithms.elite_ls_ga import EliteLSGA
from src.algorithms.baseline_pso import BaselinePSO
from src.algorithms.baseline_nsga2 import BaselineNSGA2

# =========================
# 多目标工具函数（Pareto / HV）
# =========================

def dominates_point(a, b):
    """
    最小化问题下，点 a 是否支配点 b
    a, b: (makespan, shortage)
    """
    return (
        a[0] <= b[0] and
        a[1] <= b[1] and
        (a[0] < b[0] or a[1] < b[1])
    )


def filter_nondominated(points):
    """
    输入一组二维点，返回非支配点集（最小化）
    points: List[(makespan, shortage)]
    """
    unique_points = list(set(points))
    nd = []

    for p in unique_points:
        dominated = False
        for q in unique_points:
            if p == q:
                continue
            if dominates_point(q, p):
                dominated = True
                break
        if not dominated:
            nd.append(p)

    # 按 makespan 升序，shortage 升序 排列
    nd.sort(key=lambda x: (x[0], x[1]))
    return nd


def compute_reference_point(pareto_fronts, expand_ratio=0.1):
    """
    根据某个 instance 的所有 run 的 Pareto front，自动生成 HV 的参考点
    pareto_fronts: List[List[{"makespan":..., "shortage":...}, ...]]
    """
    all_points = []
    for front in pareto_fronts:
        for sol in front:
            all_points.append((float(sol["makespan"]), float(sol["shortage"])))

    if not all_points:
        return (1.0, 1.0)

    max_m = max(p[0] for p in all_points)
    max_s = max(p[1] for p in all_points)

    ref_m = max_m * (1.0 + expand_ratio)
    ref_s = max_s * (1.0 + expand_ratio)

    return (ref_m, ref_s)


def compute_2d_hv(pareto_front, reference_point):
    """
    计算二维最小化问题的 Hypervolume
    pareto_front: List[{"makespan":..., "shortage":...}, ...]
    reference_point: (ref_makespan, ref_shortage)

    计算前，会先做一次非支配过滤。
    """
    if not pareto_front:
        return 0.0

    ref_x, ref_y = float(reference_point[0]), float(reference_point[1])

    points = [(float(sol["makespan"]), float(sol["shortage"])) for sol in pareto_front]
    nd_points = filter_nondominated(points)

    if not nd_points:
        return 0.0

    hv = 0.0
    current_y = ref_y

    # 按 makespan 升序遍历
    for x, y in nd_points:
        width = max(0.0, ref_x - x)
        height = max(0.0, current_y - y)
        hv += width * height
        current_y = min(current_y, y)

    return hv


# =========================
# 实验参数
# =========================

ALGORITHMS = {
    "EliteLSGA": EliteLSGA,
    "BaselineNSGA2": BaselineNSGA2,
}

SEEDS = list(range(1, 11))

INSTANCE_DIR = "data/instances/WIP-FMS"

RUN_RESULT_DIR = "experiments/results/runs"
SUMMARY_DIR = "experiments/results/summary"

POP_SIZE = 100
GENERATIONS = 50


# =========================
# 获取实例
# =========================

def get_instances():
    files = []

    for f in os.listdir(INSTANCE_DIR):
        if f.endswith(".json"):
            files.append(os.path.join(INSTANCE_DIR, f))

    files.sort()
    return files


# =========================
# 单次运行
# =========================

def run_once(instance_path, seed, algo_name):
    spec, operations, buffers, _ = load_instance_from_json(instance_path)

    # 双目标下，先给所有 buffer 一个默认 low_wip（如果实例里没写）
    for bid in buffers:
        if "low_wip" not in buffers[bid]:
            buffers[bid]["low_wip"] = 1

        AlgoClass = ALGORITHMS[algo_name]

        if algo_name == "BaselineNSGA2":
            search = AlgoClass(
                operations=operations,
                buffers=buffers,
                pop_size=POP_SIZE,
                n_generations=GENERATIONS,
                seed=seed
            )

            t0 = time.time()
            best = search.run(
                store_stats_init=True,
                store_stats_generations=False,
                verbose=False
            )
            runtime = time.time() - t0

            if best.stats is None:
                search.evaluate_individual(best, store_stats=True)

            pareto_front = search.get_pareto_front()

        
        elif algo_name == "EliteLSGA":
            search = AlgoClass(
                operations=operations,
                buffers=buffers,
                pop_size=POP_SIZE,
                n_generations=GENERATIONS,
                seed=seed,
                elite_ls_count=2,
                ls_max_tries=6,
                ls_blocking_threshold=3
            )

        else:
            raise ValueError(f"未知算法: {algo_name}")

        t0 = time.time()
        best = search.run(
            store_stats_init=True,
            store_stats_generations=False,
            verbose=False
        )
        runtime = time.time() - t0

        if best.stats is None:
            search.evaluate_individual(best, store_stats=True)

        pareto_front = search.get_pareto_front()

    blocking = best.stats["blocking"]["total_blocking_time"]
    shortage = best.shortage if best.shortage is not None else best.stats["shortage"]["total_shortage_area"]

    pareto_solutions = []
    for ind in pareto_front:
        if ind.stats is None:
            search.evaluate_individual(ind, store_stats=True)

        pareto_solutions.append({
            "makespan": ind.makespan,
            "shortage": ind.shortage,
            "blocking": ind.stats["blocking"]["total_blocking_time"],
            "rank": ind.rank,
            "crowding_distance": (
                "inf" if ind.crowding_distance == float("inf") else ind.crowding_distance
            ),
            "OS": ind.OS,
            "MS": ind.MS,
        })

    return {
        "rep_makespan": best.makespan,
        "rep_shortage": shortage,
        "rep_blocking": blocking,
        "runtime": runtime,
        "history_makespan": search.history_best_fitness,
        "history_shortage": getattr(search, "history_best_shortage", []),
        "pareto_size": len(pareto_solutions),
        "pareto_front": pareto_solutions,
        "hv": None,   # 先占位，后面统一计算
    }


# =========================
# 保存 run JSON
# =========================

def save_run(instance_name, algo_name, seed, result):
    algo_dir = os.path.join(RUN_RESULT_DIR, algo_name)
    os.makedirs(algo_dir, exist_ok=True)

    filename = f"{instance_name}_seed{seed}.json"
    path = os.path.join(algo_dir, filename)

    data = {
        "algorithm": algo_name,
        "instance": instance_name,
        "seed": seed,
        "parameters": {
            "pop_size": POP_SIZE,
            "generations": GENERATIONS
        },
        "representative_result": {
            "makespan": result["rep_makespan"],
            "shortage": result["rep_shortage"],
            "blocking": result["rep_blocking"],
            "runtime": result["runtime"]
        },
        "pareto_summary": {
            "pareto_size": result["pareto_size"],
            "hv": result["hv"]
        },
        "pareto_front": result["pareto_front"],
        "history": {
            "rep_makespan_per_generation": result["history_makespan"],
            "rep_shortage_per_generation": result["history_shortage"]
        }
    }

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

# =========================
# 实验主循环
# =========================

def run_experiments():
    os.makedirs(RUN_RESULT_DIR, exist_ok=True)

    instance_paths = get_instances()

    total_runs = len(instance_paths) * len(ALGORITHMS) * len(SEEDS)
    run_counter = 0

    summary = {}

    for path in instance_paths:
        instance_name = os.path.splitext(os.path.basename(path))[0]

        print("\n==============================")
        print("Instance:", instance_name)

        summary[instance_name] = {}

        # ---------------------------------
        # Phase 1: 先跑当前 instance 下所有算法、所有 seed
        # 保存到 instance_results_by_algo
        # ---------------------------------
        instance_results_by_algo = {}

        for algo_name in ALGORITHMS:
            print("Algorithm:", algo_name)
            algo_results = []

            for seed in SEEDS:
                result = run_once(path, seed, algo_name)
                algo_results.append((seed, result))
                run_counter += 1

                print(
                    f"[RUN {run_counter}/{total_runs}] "
                    f"instance={instance_name}  "
                    f"algo={algo_name}  "
                    f"seed={seed}  "
                    f"rep_makespan={result['rep_makespan']}  "
                    f"rep_shortage={result['rep_shortage']:.2f}  "
                    f"pareto_size={result['pareto_size']}  "
                    f"time={result['runtime']:.2f}s"
                )

            instance_results_by_algo[algo_name] = algo_results

        # ---------------------------------
        # Phase 2: 统一构造当前 instance 的 reference point
        # 这里必须跨“所有算法”
        # ---------------------------------
        all_pareto_fronts = []
        for algo_name in ALGORITHMS:
            for seed, result in instance_results_by_algo[algo_name]:
                all_pareto_fronts.append(result["pareto_front"])

        reference_point = compute_reference_point(all_pareto_fronts, expand_ratio=0.1)

        # ---------------------------------
        # Phase 3: 对每个算法计算 HV 并做 summary
        # ---------------------------------
        for algo_name in ALGORITHMS:
            algo_results = instance_results_by_algo[algo_name]

            rep_makespans = []
            rep_shortages = []
            pareto_sizes = []
            hvs = []

            for seed, result in algo_results:
                hv = compute_2d_hv(result["pareto_front"], reference_point)
                result["hv"] = hv

                rep_makespans.append(result["rep_makespan"])
                rep_shortages.append(result["rep_shortage"])
                pareto_sizes.append(result["pareto_size"])
                hvs.append(hv)

                save_run(instance_name, algo_name, seed, result)

            rep_makespan_avg = statistics.mean(rep_makespans)
            rep_shortage_avg = statistics.mean(rep_shortages)
            pareto_size_avg = statistics.mean(pareto_sizes)
            hv_avg = statistics.mean(hvs)
            hv_std = statistics.stdev(hvs) if len(hvs) > 1 else 0.0

            summary[instance_name][algo_name] = {
                # "rep_makespan_avg": rep_makespan_avg,
                # "rep_shortage_avg": rep_shortage_avg,
                "pareto_size_avg": pareto_size_avg,
                "hv_avg": hv_avg,
                "hv_std": hv_std,
                "reference_point": reference_point,
            }

            print(
                f"[SUMMARY] instance={instance_name}  "
                f"algo={algo_name}  "
                f"hv_avg={hv_avg:.2f}  "
                f"hv_std={hv_std:.2f}  "
                # f"rep_makespan_avg={rep_makespan_avg:.2f}  "
                # f"rep_shortage_avg={rep_shortage_avg:.2f}  "
                f"pareto_size_avg={pareto_size_avg:.2f}"
            )

    return summary

# =========================
# 保存 Excel
# =========================

def save_excel(summary):
    os.makedirs(SUMMARY_DIR, exist_ok=True)

    path = os.path.join(SUMMARY_DIR, "compare_results.xlsx")

    wb = Workbook()
    ws = wb.active

    algorithms = list(ALGORITHMS.keys())

    header = ["Instance"]

    for algo in algorithms:
        header.append(f"{algo}_hv_avg")
        header.append(f"{algo}_hv_std")
        header.append(f"{algo}_rep_makespan_avg")
        header.append(f"{algo}_rep_shortage_avg")
        header.append(f"{algo}_pareto_size_avg")

    ws.append(header)

    bold_font = Font(bold=True)

    for instance in summary:
        row = [instance]
        hv_values = []

        for algo in algorithms:
            hv_avg = summary[instance][algo]["hv_avg"]
            hv_std = summary[instance][algo]["hv_std"]
            rep_makespan_avg = summary[instance][algo]["rep_makespan_avg"]
            rep_shortage_avg = summary[instance][algo]["rep_shortage_avg"]
            pareto_size_avg = summary[instance][algo]["pareto_size_avg"]

            hv_values.append(hv_avg)

            row.append(hv_avg)
            row.append(hv_std)
            row.append(rep_makespan_avg)
            row.append(rep_shortage_avg)
            row.append(pareto_size_avg)

        ws.append(row)

        # 找最优 best makespan
        # 找最优 HV（越大越好）
        max_value = max(hv_values)
        row_id = ws.max_row
        col = 2

        for val in hv_values:
            if val == max_value:
                ws.cell(row=row_id, column=col).font = bold_font
            col += 5

    wb.save(path)

    print("\nExcel summary saved to:")
    print(path)


# =========================
# main
# =========================

if __name__ == "__main__":
    summary = run_experiments()
    save_excel(summary)